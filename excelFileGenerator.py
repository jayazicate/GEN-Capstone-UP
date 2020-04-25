from openpyxl import Workbook
import csv

# write model output to xlsx
wb = Workbook()
dest_file = "Washington.xlsx"
sheet1 = wb.active

# compile dictionary for output
fin_output = {} # license_id: {image1: risk_score, image2: risk_score, ... }
fin_output_risk_sort = {} # risk_score: {license_id: {image1: risk_score, image2: risk_score, ... }}

# store the csv data to the fin_output dictionary
with open("output.csv") as csv_file:
    reader = csv.reader(csv_file, delimiter=',')
    for row in reader:
        if row[1] != "License ID":
            image = row[0].strip()
            lic = int(row[1].strip())
            category = row[2].strip()
            output = row[3].strip()
            if category == "Storefront" or category == "Face":
                output = float(output)
            # in the dictionary already
            if lic in fin_output:
                if category in fin_output[lic]:
                    fin_output[lic][category][image] = output
                else:
                    fin_output[lic][category] = {image: output}
            elif lic not in fin_output:
                fin_output[lic] = {category: {image: output}}
            

# create the first sheet
sheet1.title = "Sorted by Image"

for id in sorted(fin_output.keys()):
    sheet1.append(("",))
    sheet1.append(("License id: " + str(id),))
    combinedScore = 0.0
    numberOf = 0
    for cat in fin_output[id].keys():
        for image in fin_output[id][cat].keys():
            output = fin_output[id][cat][image]
            sheet1.append((image, cat, output))
            if (cat == "Storefront") or (cat == "Face"):
                combinedScore += output
                numberOf += 1
    #calculate average and append to sheet, add to dictionary
    if numberOf > 0:
        average = combinedScore/numberOf
    else:
        average = -1
    sheet1.append(("", "Average Score", average))
    # add to dictionary
    if average in fin_output_risk_sort:
        fin_output_risk_sort[average][id] = fin_output[id]
    else:
        fin_output_risk_sort[average] = {id: fin_output[id]}

# create the second sheet
sheet2 = wb.create_sheet("Sorted by Risk")
sheet2.append(("License ID", "Average Risk Score"))
for avg in sorted(fin_output_risk_sort.keys(), reverse=True):
    for id in fin_output_risk_sort[avg].keys():
        sheet2.append((id, avg))

# create the third sheet
sheet3 = wb.create_sheet("More than 6 Images, by Risk")
sheet3.append(("License ID", "Average Risk Score", "Number of Images"))
for avg in sorted(fin_output_risk_sort.keys(), reverse=True):
     for id in fin_output_risk_sort[avg].keys():
        num = 0
        if "Storefront" in fin_output[id].keys():
            num += len(fin_output[id]["Storefront"].keys())
        if "Face" in fin_output[id].keys():
            num += len(fin_output[id]["Face"].keys())
        if num >= 7:
            sheet3.append((id, avg, num))

wb.save(filename = dest_file)
